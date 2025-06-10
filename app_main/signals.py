from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.db.models.signals import post_save
from django.dispatch import receiver

from openpyxl import load_workbook

from .models import (
    ExcelDocument,
    DeliveryBatch,
    Product,
    SupplierExcelDocument,
    Supplier
)


@receiver(post_save, sender=ExcelDocument)
def populate_product_document(sender, instance, created, **kwargs):
    file = instance.document
    wb = load_workbook(file)
    sheet = wb.active

    try:
        title = sheet.cell(row=1, column=1).value or ""
        manifest_register_number = sheet.cell(row=2, column=1).value.split("№")[-1]
        total_products = sheet.cell(row=1, column=2).value.split()[-1]
        total_recipients = sheet.cell(row=2, column=2).value.split()[-1]
        sender_name = sheet.cell(row=3, column=1).value.lower().split("отправитель")[-1].strip().upper()
        total_weight = sheet.cell(row=3, column=2).value
        total_price = sheet.cell(row=4, column=2).value
        send_date = datetime.strptime(sheet.cell(row=4, column=1).value.split()[-1], "%d.%m.%Y").date()
    except Exception as e:
        print(f"[ERROR: Product Sheet] Failed parsing batch metadata: {e}")
        return

    delivery_batch, _ = DeliveryBatch.objects.get_or_create(
        title=title,
        manifest_register_number=manifest_register_number.strip() if manifest_register_number else manifest_register_number,
        total_products=total_products,
        total_recipients=total_recipients,
        sender_name=sender_name,
        total_weight=total_weight,
        total_price=total_price,
        send_date=send_date,
    )

    for row in sheet.iter_rows(min_row=6, values_only=True):
        if not row[0]:
            break

        try:
            Product.objects.get_or_create(
                delivery_batch=delivery_batch,
                product_id=row[0],
                invoice=row[1],
                awb=row[2],
                sender_number=row[3] if manifest_register_number.strip()[0] == "2" else "",
                sender_id=row[4] if manifest_register_number.strip()[0] == "2" else "",
                sticker=row[3] if manifest_register_number.strip()[0] != "2" else "",
                product_name=row[4 if manifest_register_number.strip()[0] != "2" else 5],
                netto=row[6],
                brutto=row[7],
                quantity=row[8],
                price=row[9],
                currency=row[10 if manifest_register_number.strip()[0] != "2" else 11],
                tn_ved=row[11 if manifest_register_number.strip()[0] != "2" else 12],
                qr_code=row[12] if manifest_register_number.strip()[0] == "2" else "",
                shk=row[12] if manifest_register_number.strip()[0] != "2" else "",
                recipient_fullname=row[13].title() if manifest_register_number.strip()[0] != "2" else row[15].title(),
                recipient_passport=row[14] if manifest_register_number.strip()[0] != "2" else row[16],
                recipient_pinfl=row[15] if manifest_register_number.strip()[0] != "2" else row[17],
                recipient_address=row[16] if manifest_register_number.strip()[0] != "2" else row[18],
                recipient_phonenumber=row[17] if manifest_register_number.strip()[0] != "2" else row[19],
                box_number=row[18] if manifest_register_number.strip()[0] != "2" else row[20],
                recipient_birthdate=row[19] if manifest_register_number.strip()[0] != "2" else row[21]
            )
        except Exception as e:
            print(f"[ERROR: Product Row] Failed saving product row: {e}")


@receiver(post_save, sender=SupplierExcelDocument)
def populate_supplier_document(sender, instance, created, **kwargs):
    file = instance.document
    wb = load_workbook(file)

    sheet_name = "FedEx"
    if sheet_name not in wb.sheetnames:
        print(f"[WARNING] Sheet '{sheet_name}' not found in the Excel file.")
        return

    sheet = wb[sheet_name]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            break  # Stop if the "date" column is empty (row is likely invalid)

        try:
            weight_val = Decimal(row[7]) if row[7] else None
        except (InvalidOperation, TypeError):
            weight_val = None

        try:
            supplier, created = Supplier.objects.get_or_create(
                date=row[1],
                sender=row[2].strip().title() if isinstance(row[2], str) else None,
                number=row[3],
                country=row[4].strip().title() if isinstance(row[4], str) else None,
                zone=row[5].strip().upper() if isinstance(row[5], str) else None,
                what_is_inside=row[6].strip().upper() if isinstance(row[6], str) else None,
                weight=weight_val,
                payment_type=row[8].strip().upper() if isinstance(row[8], str) else None,
                price=row[9],
                additional_percent=row[10].replace("%", "") if isinstance(row[10], str) else "",
            )
        except Exception as e:
            print(f"[ERROR: Supplier Row] Failed saving supplier row: {e}")
